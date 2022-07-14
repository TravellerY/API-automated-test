# coding=utf-8
import os
import sys

os_path = os.path.dirname(os.path.dirname(__file__))
sys.path.append(os_path)

import openpyxl
import json


class HandleExcel(object):
    def __init__(self, filepath=None, index=None):
        """
        构造函数
        :param filepath: 文件的路径
        :param index: xlsx的下标
        """
        if filepath is None:
            self.file_name = os.path.join(os.path.dirname(os.path.dirname(__file__)), r'config/APIcase.xlsx')
        else:
            self.file_name = filepath
        self.load_e = self.load_excel(self.file_name)
        if index is None:
            self.index = 0
        else:
            self.index = index

    def load_excel(self, file_name):
        """
        加载Excel文件
        """
        open_excel = openpyxl.load_workbook(file_name)
        return open_excel

    def get_sheet_data(self):
        """
        获取sheet中的值
        :return:
        """
        sheet_name = self.load_e.sheetnames
        data = self.load_e[sheet_name[self.index]]
        return data

    def get_rows(self):
        """
        获取行数
        :return:返回总行数
        """
        rows = self.get_sheet_data().max_row
        if rows:
            return rows
        else:
            return None

    def get_row_value(self, row):
        """
        获取某一行的内容
        :param row: 需获取值的行号
        :return:
        """
        row_list = []
        data = self.get_sheet_data()[row]
        if data:
            for i in data:
                row_list.append(i.value)
            return row_list
        return None

    def get_cell_value(self, row, cols):
        """
        获取某一单元格的内容
        :param row: 该单元格的行号
        :param cols: 该单元格的列号
        :return:
        """
        cell_value = self.get_sheet_data().cell(row, cols).value
        return cell_value

    def write_value(self, row, cols, value):
        """
        将数据写入到文件中
        :param row: 写入数据的行号
        :param cols: 写入数据的列号
        :param value: 需写入的数据
        :return:
        """
        if isinstance(value, dict):
            values = json.dumps(value)
        else:
            values = value
        wr = self.get_sheet_data()
        wr.cell(row, column=cols, value=values)
        self.load_e.save(self.file_name)

    def get_columns_value(self, key):
        """
        获取某一列的内容
        :param key:
        :return:
        """
        data_list = []
        data = self.get_sheet_data()[key]
        if data:
            for i in data:
                data_list.append(i.value)
            return data_list
        return None

    def get_rows_data(self):
        """
        获取所有行的内容，并将内容以list的形式输出
        :return:
        """
        rows_list = []
        rows = self.get_rows()
        if rows:
            for i in range(2, rows + 1):
                value = self.get_row_value(i)
                value.append(i)
                rows_list.append(value)
            self.load_e.close()
            return rows_list

# if __name__ == '__main__':
#     handle_e = HandleExcel()
#     # res = handle_e.get_rows_data()
#     res = handle_e.get_cell_value(2, 0)
#     print(res)
#
